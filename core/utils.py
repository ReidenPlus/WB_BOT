from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
# ВАЖНО: Импортируем модели, иначе Python не поймет, с чем сравнивать
from .models import Order, WithdrawalRequest

def export_to_excel(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Export Data"

    # Проверяем, какую модель мы выгружаем
    if queryset.model == Order:
        headers = [
            'ID', 'Пользователь', 'Артикул', 'Товар', 'Цена WB', '% Кэшбэка', 
            'Статус', 'Дата', 'Реквизиты', 'Скрин Заказа', 'Скрин Чека', 'Номер чека'
        ]
    elif queryset.model == WithdrawalRequest:
        headers = ['ID', 'Пользователь', 'Сумма', 'Реквизиты', 'Статус', 'Дата']
    else:
        headers = [field.name for field in modeladmin.model._meta.fields]

    ws.append(headers)
    
    # Жирный шрифт для шапки
    for cell in ws[1]: 
        cell.font = Font(bold=True)

    for obj in queryset:
        row = []
        
        if queryset.model == Order:
            u_name = str(obj.user) if obj.user else "Нет"
            details = obj.user.payment_details if obj.user and obj.user.payment_details else "Нет реквизитов"
            
            p_art = obj.product.article if obj.product else "-"
            p_name = obj.product.name if obj.product else "-"
            p_price = obj.product.wb_price if obj.product else 0
            p_perc = obj.product.cashback_percent if obj.product else 0
            
            s1 = obj.screenshot.url if obj.screenshot else "-"
            s2 = obj.receipt_screenshot.url if obj.receipt_screenshot else "-"
            check_num = obj.check_number if obj.check_number else "-"
            
            date_str = obj.created_at.strftime("%d.%m.%Y %H:%M")
            
            row = [
                obj.id, u_name, p_art, p_name, p_price, f"{p_perc}%", 
                obj.get_status_display(), date_str, details, s1, s2, check_num
            ]
        
        elif queryset.model == WithdrawalRequest:
            row = [
                obj.id, str(obj.user), obj.amount, obj.phone_number, 
                obj.get_status_display(), obj.created_at.strftime("%d.%m.%Y %H:%M")
            ]
        
        else:
            for field in headers:
                val = getattr(obj, field, "-")
                row.append(str(val))
        
        ws.append(row)

    # Авто-ширина
    for column_cells in ws.columns:
        length = max(len(str(cell.value) or "") for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={queryset.model._meta.model_name}_report.xlsx'
    wb.save(response)
    return response

# Настраиваем название действия здесь же
export_to_excel.short_description = "Скачать Excel отчет (.xlsx)"