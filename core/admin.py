from django.contrib import admin
from django.utils.html import format_html
from django.http import HttpResponse
from .models import TelegramUser, Product, Order, WithdrawalRequest, CartItem, ProductImage
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# --- –£–ú–ù–´–ô –≠–ö–°–ü–û–†–¢ –í EXCEL ---
def export_to_excel(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Export Data"

    # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –∫–∞–∫—É—é –º–æ–¥–µ–ª—å –º—ã –≤—ã–≥—Ä—É–∂–∞–µ–º, —á—Ç–æ–±—ã –¥–∞—Ç—å –ø—Ä–∞–≤–∏–ª—å–Ω—ã–µ –∑–∞–≥–æ–ª–æ–≤–∫–∏
    if queryset.model == Order:
        # –ó–ê–ì–û–õ–û–í–ö–ò –î–õ–Ø –ó–ê–ö–ê–ó–û–í
        headers = [
            'ID', '–ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å', '–ê—Ä—Ç–∏–∫—É–ª', '–¢–æ–≤–∞—Ä', '–¶–µ–Ω–∞ WB', '% –ö—ç—à–±—ç–∫–∞', 
            '–°—Ç–∞—Ç—É—Å', '–î–∞—Ç–∞', '–†–µ–∫–≤–∏–∑–∏—Ç—ã', '–°–∫—Ä–∏–Ω –ó–∞–∫–∞–∑–∞', '–°–∫—Ä–∏–Ω –ß–µ–∫–∞', '–ù–æ–º–µ—Ä —á–µ–∫–∞'
        ]
    elif queryset.model == WithdrawalRequest:
        headers = ['ID', '–ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å', '–°—É–º–º–∞', '–†–µ–∫–≤–∏–∑–∏—Ç—ã', '–°—Ç–∞—Ç—É—Å', '–î–∞—Ç–∞']
    else:
        headers = [field.name for field in modeladmin.model._meta.fields]

    ws.append(headers)
    
    for cell in ws[1]: 
        cell.font = Font(bold=True)

    for obj in queryset:
        row = []
        
        if queryset.model == Order:
            u_name = str(obj.user) if obj.user else "–ù–µ—Ç"
            details = obj.user.payment_details if obj.user and obj.user.payment_details else "–ù–µ—Ç —Ä–µ–∫–≤–∏–∑–∏—Ç–æ–≤"
            
            p_art = obj.product.article if obj.product else "-"
            p_name = obj.product.name if obj.product else "-"
            p_price = obj.product.wb_price if obj.product else 0
            p_perc = obj.product.cashback_percent if obj.product else 0
            
            s1 = obj.screenshot.url if obj.screenshot else "-"
            s2 = obj.receipt_screenshot.url if obj.receipt_screenshot else "-"
            
            check_num = obj.check_number if obj.check_number else "-"
            
            date_str = obj.created_at.strftime("%d.%m.%Y %H:%M")
            
            row = [
                obj.id, u_name, p_art, p_name, p_price, f"{p_perc}%", 
                obj.get_status_display(), date_str, details, s1, s2, check_num
            ]
        
        elif queryset.model == WithdrawalRequest:
            row = [
                obj.id, str(obj.user), obj.amount, obj.phone_number, 
                obj.get_status_display(), obj.created_at.strftime("%d.%m.%Y %H:%M")
            ]
        
        else:
            for field in headers:
                val = getattr(obj, field, "-")
                row.append(str(val))
        
        ws.append(row)
        
    for column_cells in ws.columns:
        length = max(len(str(cell.value) or "") for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={queryset.model._meta.model_name}_report.xlsx'
    wb.save(response)
    return response

export_to_excel.short_description = "–°–∫–∞—á–∞—Ç—å Excel –æ—Ç—á–µ—Ç (.xlsx)"

@admin.action(description="üì¶ –í –ê–†–•–ò–í (–°–∫—Ä—ã—Ç—å)")
def move_to_archive(modeladmin, request, queryset):
    queryset.update(is_archived=True)

@admin.action(description="‚ôªÔ∏è –í–û–°–°–¢–ê–ù–û–í–ò–¢–¨ –∏–∑ –∞—Ä—Ö–∏–≤–∞")
def restore_from_archive(modeladmin, request, queryset):
    queryset.update(is_archived=False)

class ProductImageInline(admin.TabularInline):
    model = ProductImage
    extra = 1

@admin.register(Product)
class ProductAdmin(admin.ModelAdmin):
    list_display = ('name', 'article', 'wb_price', 'cashback_percent', 'active', 'is_archived')
    list_filter = ('active', 'is_archived')
    search_fields = ('name', 'article')
    actions = [move_to_archive, restore_from_archive]
    inlines = [ProductImageInline]

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if 'is_archived__exact' not in request.GET:
            return qs.filter(is_archived=False)
        return qs

@admin.register(Order)
class OrderAdmin(admin.ModelAdmin):
    list_display = ('id', 'user', 'product_info', 'status', 'calc_cashback', 'check_number', 'created_at', 'view_screens')
    list_filter = ('status', 'is_archived', 'created_at')
    search_fields = ('user__username', 'user__telegram_id', 'product__article', 'check_number')
    list_select_related = ('user', 'product')
    
    actions = ['set_received', 'set_approved', 'set_rejected', move_to_archive, restore_from_archive, export_to_excel]

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if 'is_archived__exact' not in request.GET:
            return qs.filter(is_archived=False)
        return qs

    @admin.display(description="–¢–æ–≤–∞—Ä")
    def product_info(self, obj):
        return f"{obj.product.name} (–ê—Ä—Ç: {obj.product.article})"

    @admin.display(description="–ö –≤—ã–ø–ª–∞—Ç–µ")
    def calc_cashback(self, obj):
        if not obj.product: return "0 ‚ÇΩ"
        amount = int(obj.product.wb_price * obj.product.cashback_percent / 100)
        return f"{amount} ‚ÇΩ"

    @admin.display(description="–°–∫—Ä–∏–Ω—à–æ—Ç—ã")
    def view_screens(self, obj):
        html = ""
        if obj.screenshot:
            html += format_html('<a href="{}" target="_blank">üñºÔ∏è –õ–ö</a> ', obj.screenshot.url)
        if obj.receipt_screenshot:
            html += format_html('<br><a href="{}" target="_blank">üßæ –ß–µ–∫</a>', obj.receipt_screenshot.url)
        return format_html(html) if html else "-"

    @admin.action(description="–°—Ç–∞—Ç—É—Å -> ‚úÖ –ü–æ–ª—É—á–µ–Ω")
    def set_received(self, request, queryset):
        queryset.update(status='received')

    @admin.action(description="–°—Ç–∞—Ç—É—Å -> ‚ùå –û—Ç–∫–ª–æ–Ω–µ–Ω–æ")
    def set_rejected(self, request, queryset):
        queryset.update(status='rejected')

    @admin.action(description="–°—Ç–∞—Ç—É—Å -> üí∞ –í—ã–ø–ª–∞—á–µ–Ω–æ")
    def set_approved(self, request, queryset):
        count = 0
        for order in queryset:
            if order.status != 'approved':
                order.status = 'approved'
                cash = (order.product.wb_price * order.product.cashback_percent) / 100
                order.user.balance += cash
                order.user.save()
                order.save()
                count += 1
        self.message_user(request, f"–í—ã–ø–ª–∞—á–µ–Ω–æ –∑–∞–∫–∞–∑–æ–≤: {count}")

@admin.register(WithdrawalRequest)
class WithdrawalAdmin(admin.ModelAdmin):
    list_display = ('user', 'amount', 'phone_number', 'status')
    actions = [export_to_excel]

admin.site.register(TelegramUser)
admin.site.register(CartItem)
